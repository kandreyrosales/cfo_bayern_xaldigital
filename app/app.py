import os
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import boto3
import jwt
from datetime import datetime
from functools import wraps
import psycopg2

from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

app = Flask(__name__)

app.secret_key = 'xaldigitalcfobayer!'
AWS_REGION_PREDICTIA = os.getenv("region_aws", 'us-east-1')
bucket_name = os.getenv("bucket_name")
accessKeyId = os.getenv("accessKeyId")
secretAccessKey = os.getenv("secretAccessKey")
CLIENT_ID_COGNITO =os.getenv("client_id")
USER_POOL_ID_COGNITO =os.getenv("user_pool")

# boto3 clients
cognito_client = boto3.client('cognito-idp', region_name=AWS_REGION_PREDICTIA, aws_access_key_id=accessKeyId, aws_secret_access_key=secretAccessKey)
lambda_client = boto3.client('lambda', region_name=AWS_REGION_PREDICTIA, aws_access_key_id=accessKeyId, aws_secret_access_key=secretAccessKey)
s3_client = boto3.client('s3', region_name=AWS_REGION_PREDICTIA, aws_access_key_id=accessKeyId, aws_secret_access_key=secretAccessKey)

db_host = os.getenv("db_endpoint")
db_name = "postgres"
db_user = "cfo_user"
db_password = os.getenv("password_db")

def authenticate_user(username, password):
    try:
        response = cognito_client.admin_initiate_auth(
            AuthFlow='ADMIN_NO_SRP_AUTH',
            AuthParameters={
                'USERNAME': username,
                'PASSWORD': password
            },
            ClientId=CLIENT_ID_COGNITO,
            UserPoolId=USER_POOL_ID_COGNITO,
            ClientMetadata={
                'username': username,
                'password': password,
            }
        )
        return response
    except cognito_client.exceptions.NotAuthorizedException as e:
        # Handle invalid credentials
        return {"reason": "Credenciales Inválidas"}
    except cognito_client.exceptions.ResourceNotFoundException as e:
        # Handle invalid credentials
        return {"reason": "Recurso No Encontrado"}
    except cognito_client.exceptions.UserNotFoundException as e:
        # Handle invalid credentials
        return {"reason": "Usuario No Encontrado"}
    except cognito_client.exceptions.UserNotConfirmedException as e:
        # Handle invalid credentials
        return {"reason": "Usuario No Confirmado"}
    except Exception as e:
        # Handle other errors
        return {"reason": "Error general. Por favor contactar al administrador"}


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
        Accessing with Cognito using username and password.
        After login is redirected to reset password and login again
    """
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            return render_template('login/login.html',
                                   error="Nombre de usuario y Contraseña obligatorios")

        cognito_response = authenticate_user(username, password)

        reason = cognito_response.get("reason")
        if reason == "Usuario No Confirmado":
            return render_template('login/confirm_account_code.html', email=username)
        elif reason is not None:
            return render_template('login/login.html', error=reason)

        auth_result = cognito_response.get("AuthenticationResult")
        if not auth_result:
            return render_template('login/login.html', error=cognito_response)

        session['access_token'] = auth_result.get('AccessToken')
        session['id_token'] = auth_result.get('IdToken')
        return redirect(url_for('index'))
    else:
        return render_template(
            'login/login.html',
            accessKeyId=accessKeyId,
            secretAccessKey=secretAccessKey
        )

@app.route('/registro', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        corporate_title = request.form['corp_title']
        contact_number = request.form['contact_number']
        user_id = request.form['worker_custom_id']
        user_attributes = [
            {
                'Name': 'custom:corporate_title',
                'Value': corporate_title
            },
            {
                'Name': 'custom:contact_number',
                'Value': contact_number
            },
            {
                'Name': 'custom:user_id',
                'Value': user_id
            }
        ]
        try:
            response = cognito_client.sign_up(
                ClientId=CLIENT_ID_COGNITO,
                Username=username,
                Password=password,
                UserAttributes=user_attributes
            )
            return render_template('login/confirm_account_code.html', email=username)
        except cognito_client.exceptions.NotAuthorizedException as e:
            # Handle authentication failure
            return render_template(
                'login/signup.html',
                error="Usuario no Autorizado para ejecutar esta acción")
        except cognito_client.exceptions.UsernameExistsException:
            return render_template(
                'login/signup.html',
                error="Ya existe una cuenta asociada a este correo")
        except cognito_client.exceptions.InvalidPasswordException:
            return render_template(
                'login/signup.html',
                error="Crea una contraseña de al menos 8 dígitos, más segura, usando al menos una letra mayúscula, "
                      "un número y un carácter especial")
        except Exception as e:
            return render_template(
                'login/signup.html',
                error=f"Ha ocurrido el siguiente error: {e}")
    else:
        return render_template('login/signup.html')

@app.route('/confirmar_cuenta', methods=['GET', 'POST'])
def confirm_account_code():
    email = request.form['email_not_confirmed']
    email_code = request.form['custom_code']
    if request.method == "POST":
        try:
            cognito_client.confirm_sign_up(
                ClientId=CLIENT_ID_COGNITO,
                Username=email,
                ConfirmationCode=email_code
            )
            return render_template('login/login.html')
        except cognito_client.exceptions.ExpiredCodeException as e:
            return render_template(
                'login/confirm_account_code.html',
                error="El código enviado a su correo ha expirado", email=email)
        except cognito_client.exceptions.CodeMismatchException as e:
            return render_template(
                'login/confirm_account_code.html',
                error="El código no es válido", email=email)
        except cognito_client.exceptions.TooManyFailedAttemptsException as e:
            return render_template(
                'login/confirm_account_code.html',
                error="Máximo de intentos superados para validar la cuenta", email=email)
        except cognito_client.exceptions.UserNotFoundException as e:
            return render_template(
                'login/confirm_account_code.html',
                error="Error: Usuario no Encontrado o Eliminado", email=email)
    else:
        return render_template('login/confirm_account_code.html', email=email)

@app.route('/logout')
def logout():
    # Clear the session data
    session.clear()
    return redirect(url_for('login'))

def token_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = session.get("access_token")
        if not token:
            return render_template('login/login.html')
        try:
            decoded_token = jwt.decode(token, options={"verify_signature": False})  # Decode the token without verifying signature
            expiration_time = datetime.utcfromtimestamp(decoded_token['exp'])
            current_time = datetime.utcnow()
            if expiration_time > current_time:
                return f(*args, **kwargs)
            else:
                return render_template('login/login.html',
                                       error="Sesión Expirada. Ingrese sus datos de nuevo")
        except jwt.ExpiredSignatureError:
            return render_template('login/login.html',
                                   error="Sesión Expirada. Ingrese sus datos de nuevo")
    return decorated_function

@app.route('/olvido_contrasena', methods=['GET', 'POST'])
def forgot_password():
    if request.method == "POST":
        email = request.form['email_forgot_password']
        password = request.form['password']
        custom_code = request.form['custom_code']
        try:
            cognito_client.confirm_forgot_password(
                ClientId=CLIENT_ID_COGNITO,
                Username=email,
                ConfirmationCode=custom_code,
                Password=password,
            )
            return render_template('login/login.html')
        except cognito_client.exceptions.UserNotFoundException as e:
            return render_template('login/reset_password.html',
                                   error="Usuario No Encontrado o Eliminado.", email=email)
        except cognito_client.exceptions.InvalidPasswordException as e:
            return render_template('login/reset_password.html',
                                   error="Usuario No Encontrado o Eliminado.", email=email)
        except cognito_client.exceptions.CodeMismatchException as e:
            return render_template('login/reset_password.html',
                                   error="Ha ocurrido un problema al enviar el código para asignar una nueva contraseña. "
                                         "Intenta de nuevo.",
                                   email=email)
        except Exception as e:
            return render_template('login/reset_password.html',
                                   error=f"Error: {e}", email=email)
    else:
        return render_template('login/reset_password.html')

@app.route('/enviar_link_contrasena', methods=['GET', 'POST'])
def send_reset_password_link():
    if request.method == "POST":
        email = request.form['email_forgot_password']
        try:
            cognito_client.forgot_password(
                ClientId=CLIENT_ID_COGNITO,
                Username=email
            )
            return render_template('login/reset_password.html')

        except cognito_client.exceptions.UserNotFoundException as e:
            return render_template('login/send_reset_password_link.html',
                                   error="Usuario No Encontrado o Eliminado.", email=email)
        except cognito_client.exceptions.CodeDeliveryFailureException as e:
            return render_template('login/send_reset_password_link.html',
                                   error="Ha ocurrido un problema al enviar el código para asignar una nueva contraseña. "
                                         "Intenta de nuevo.",
                                   email=email)
        except Exception as e:
            return render_template('login/send_reset_password_link.html',
                                   error=f"Error: {e}", email=email)
    else:
        return render_template('login/send_reset_password_link.html')

@app.route('/', methods=["GET"])
# @token_required
def index():
    # try:
    #     cognito_client.get_user(AccessToken=session.get("access_token"))
    # except cognito_client.exceptions.UserNotFoundException as e:
    #     return redirect(url_for('logout'))
    transactions_labels = ["Totales", "Exitosas", "Fallidas", "Fallidas sin CFDI", "Falta emitir Complemento"]
    transactions_data = [35, 26, 6, 1, 3]

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    rfc = request.args.get("rfc")
    customer = request.args.get("customer")

    where_statement_sql, filters_present = generate_filter_sql(
        start_date=start_date,
        end_date=end_date,
        rfc=rfc,
        customer=customer
    )
    if not filters_present:
        where_statement_sql = ""


    # Validador IVA table data
    conn, cur = connection_db()
    query_validador_ivas = f"""
        select iva_cobrado_sat,  iva, validar_ivas_validador_iva
        from dashboard {where_statement_sql} limit 5
    """
    query_validador_ivas_rows = get_query_rows(
        conn=conn,
        cur=cur,
        query=query_validador_ivas)

    # Validador IEPS table data
    conn, cur = connection_db()
    query_validador_ieps = f"""
        select ieps_cobrado_sat,  ieps, validador_ieps_validador_iva
        from dashboard {where_statement_sql} limit 5
    """
    query_validador_ieps_rows = get_query_rows(
        conn=conn,
        cur=cur,
        query=query_validador_ieps)

    # # Cliente con mayor variacion Chart Data
    # conn, cur = connection_db()
    # query_clientes_mayor_variacion = f"""
    # select cliente,
    #    CASE
    #         WHEN SUM(total_variacion_validador_iva) > 0
    #             THEN round(SUM(total_variacion_validador_iva), 2)
    #         ELSE 0
    #    END AS total_variacion_por_cliente
    # from conciliaciones_raw_data
    # {where_statement_sql}
    # group by cliente
    # order by total_variacion_por_cliente DESC"""
    # query_clientes_mayor_variacion_rows = get_query_rows(
    #     conn=conn,
    #     cur=cur,
    #     query=query_clientes_mayor_variacion)
    #
    # variation_labels = []
    # variation_data = []
    # if query_clientes_mayor_variacion_rows:
    #     variation_labels = [row[0] for row in query_clientes_mayor_variacion_rows]
    #     variation_data = [row[1] for row in query_clientes_mayor_variacion_rows]

    variation_labels = ["MONSANTO COMERCIAL", "Bayer AG Crop Science", "TIENDAS SORIANA",
                        "2022 ENVIRONMENTAL SCIENCE", "Bayer HealthCare LLC Consumer Care Division"]
    variation_data = [93.79, 24.66, 2.45, 1.38, 0.58]

    conn, cur = connection_db()
    transaction_sum_query_result = get_query_rows(
        conn=conn,
        cur=cur,
        query=f"""select count(*) from cfdi_ingreso;"""
    )[0][0]

    conn, cur = connection_db()
    ingresos_totales_query = """select SUM(total_cfdi) FROM cfdi_ingreso;"""
    ingresos_totales_query_result = get_query_rows(
        conn=conn,
        cur=cur,
        query=ingresos_totales_query
    )[0][0]


    conn, cur = connection_db()
    ingresos_data_chart_query = """
        select cliente, sum(depositos) as total from conciliaciones_raw_data
        group by cliente order by total DESC;"""
    ingresos_data_chart_query_result = get_query_rows(
        conn=conn,
        cur=cur,
        query=ingresos_data_chart_query
    )
    ingresos_chart_labels = [row[0] for row in ingresos_data_chart_query_result]
    ingresos_chart_data = [row[1] for row in ingresos_data_chart_query_result]

    return render_template(
        'index.html',
        variation_data=variation_data,
        variation_labels=variation_labels,
        transactions_labels=transactions_labels,
        transactions_data=transactions_data,
        validador_iva_rows=query_validador_ivas_rows,
        validador_ieps_rows=query_validador_ieps_rows,
        transactions_count = transaction_sum_query_result,
        ingresos_totales = ingresos_totales_query_result,
        ingresos_chart_data = ingresos_chart_data,
        ingresos_chart_labels = ingresos_chart_labels
    )

@app.route('/get_rfc_list', methods=["GET"])
def get_rfc_list():
    """
    Getting the RFC list from the Database
    """
    conn, cur = connection_db()
    query_uuid = """select DISTINCT (rfc) from conciliaciones ;"""
    data_from_db = get_query_rows(cur=cur, conn=conn, query=query_uuid)
    new_item = ("Todos", )
    try:
        return [new_item]+data_from_db
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/get_customer_name_list', methods=["GET"])
def get_customer_name_list():
    """
    Getting customer names list from the Database
    """
    conn, cur = connection_db()
    query_uuid = """select DISTINCT (cliente) from conciliaciones ;"""
    data_from_db = get_query_rows(cur=cur, conn=conn, query=query_uuid)
    new_item = ("Todos", )
    try:
        return [new_item]+data_from_db
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/conciliaciones')
@token_required
def reconciliations_data_cfo():
    # try:
    #     cognito_client.get_user(AccessToken=session.get("access_token"))
    # except cognito_client.exceptions.UserNotFoundException as e:
    #     return redirect(url_for('logout'))
    conn, cur = connection_db()
    query_conciliations_view = """select * from conciliaciones order by cliente, fecha;"""
    rows_data_view = get_query_rows(cur=cur, conn=conn, query=query_conciliations_view)
    return render_template(
        'reconciliations_data_cfo.html',
        initial_data_for_table=rows_data_view,
    )

def generate_filter_sql(start_date, end_date, rfc, customer):
    """
    Smart query for Where statement using variables from the page
    """
    where_query = """WHERE """
    query_date_range = f"""fecha BETWEEN '{start_date}' and '{end_date}' """
    query_rfc = f""" rfc='{rfc}' """
    query_customer = f""" cliente='{customer}' """

    field_flag = False
    if start_date and end_date:
        field_flag = True
        where_query += f"{query_date_range}"
    if rfc and rfc != 'all':
        if field_flag:
            where_query += f" AND {query_rfc}"
        else:
            where_query += f"{query_rfc}"
        field_flag = True
    if customer and customer != 'all':
        if field_flag:
            where_query += f" AND {query_customer}"
        else:
            where_query += f"{query_customer}"
            field_flag = True
    if field_flag:
        return where_query, True
    return where_query, False

@app.route('/get_filtered_data_conciliations', methods=["GET"])
def get_filtered_data_conciliations():
    conn, cur = connection_db()

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    rfc = request.args.get("rfc")
    customer = request.args.get("customer")

    select_table_fields = """SELECT rfc, CAST(factura_bayer AS BIGINT), cliente, transaccion, to_char(fecha, 'DD/MM/YYYY'), estado, 
            uuid, subtotal, iva, ieps, total,
            depositos, nombre_del_banco, to_char(fecha_deposito, 'DD/MM/YYYY'), validador_aplicacion_pagos, 
            CAST(document_number_sap as BIGINT), CAST(clearing_document_sap AS BIGINT), subtotal_sap, iva_sap, ieps_sap, total_aplicacion_sap, 
            uuid_relacionado, subtotal_sat, iva_cobrado_sat, ieps_cobrado_sat, total_aplicacion_sat, 
            validador_subtotal_validador_iva, validar_ivas_validador_iva,
            validador_ieps_validador_iva, total_variacion_validador_iva"""

    where_statement_sql, filters_present = generate_filter_sql(
        start_date=start_date,
        end_date=end_date,
        rfc=rfc,
        customer=customer
    )
    if filters_present:
        query_conciliations_view_filtered = f"""
            {select_table_fields} from conciliaciones {where_statement_sql} order by cliente, fecha
        """
    else:
        query_conciliations_view_filtered = f"""
                    {select_table_fields} from conciliaciones order by cliente, fecha
                """
    result = get_query_rows(
        cur=cur,
        conn=conn,
        query=query_conciliations_view_filtered
    )
    return jsonify(result)


@app.route('/vista_subir_archivo')
def uploadfile():
    return render_template('uploadfile.html')

def custom_values_for_insert(data_sheet, max_col: int):
    # Prepare data for bulk insertion
    data_collection = []
    for row in data_sheet.iter_rows(min_row=2, max_col=max_col):  # Start from the second row (data)
        if row[0].value == "" or row[0].value is None:
            continue
        data_collection.append(f"""{tuple(str(cell.value).replace("'","") if cell.value is not None else '' for cell in row)}""")
    return data_collection

def connection_db():
    try:
        conn = psycopg2.connect(
            dbname=db_name,
            user=db_user,
            password=db_password,
            host=db_host.split(":")[0],
            port=5432)
        cur = conn.cursor()
        return conn, cur
    except Exception as e:
        print(f"Error connecting to database: {e}")
        exit()

def get_query_rows(cur, conn, query):
    try:
        cur.execute(query)
        rows = cur.fetchall()  # Fetch all rows
        # Print fetched rows (optional)
        return rows
    except (Exception, psycopg2.Error) as error:
        print("Error while fetching data from PostgreSQL", error)
    finally:
        # Close the connection
        if conn:
            cur.close()
            conn.close()
        print("PostgreSQL connection is closed")
def execute_query(cur, conn, query):
    # Execute bulk insert using executemany
    try:
        cur.execute(query)
        conn.commit()
        print(f"Data inserted successfully into table")
    except Exception as e:
        print(f"Error inserting data: {e}")
        conn.rollback()  # Rollback changes in case of errors

@app.route('/subir_archivo', methods=['POST'])
def subir_archivo():
    if 'archivo' not in request.files:
        return 'No se ha enviado ningún archivo'
    archivo = request.files['archivo']
    if archivo.filename == '':
        return 'No se ha seleccionado ningún archivo'
    # Verificar si el archivo tiene la extensión permitida
    if not archivo.filename.endswith('.xlsx'):
        return 'La extensión del archivo no está permitida. Se permiten solo archivos .xlsx'
    stream = BytesIO(archivo.read())
    workbook = load_workbook(stream)
    cfdi_ingresos_sheet = workbook.worksheets[0]
    cfdi_complemento_sheet = workbook.worksheets[1]
    iva_cobrado_sheet = workbook.worksheets[2]
    analisis_iva_cobrado_sheet = workbook.worksheets[3]
    # Prepare bulk insert query template
    try:
        column_names_ingresos = [
            "version",
            "estado",
            "tipo_comprobante",
            "rfc",
            "nombre",
            "fecha_emision",
            "folio_interno",
            "uuid_fiscal",
            "producto_servicio_conforme_sat",
            "concepto_del_cfdi",
            "moneda",
            "metodo_de_pago",
            "subtotal",
            "descuento",
            "iva",
            "ieps",
            "isr_retenido",
            "iva_retenido",
            "total_cfdi",
            "tipo_de_relacion",
            "cfdi_relacionado",]
        column_names_str_ingresos = ", ".join(column_names_ingresos)
        column_names_complemento = [
            "estado_sat",
            "version_pagos",
            "uuid_complemento",
            "fecha_timbrado",
            "fecha_emision",
            "folio",
            "serie",
            "subtotal",
            "moneda",
            "total",
            "lugar_expedicion",
            "rfc_emisor",
            "nombre_emisor",
            "rfc_receptor",
            "nombre_receptor",
            "uso_cfdi",
            "clave_prod_serv",
            "descripcion",
            "fecha_de_pago",
            "forma_de_pago",
            "moneda_p",
            "tipo_cambio_p",
            "monto",
            "numero_operacion",
            "importe_pagado_mo",
            "id_documento",
            "serie_dr",
            "folio_dr",
            "moneda_dr",
            "num_parcialidad",
            "imp_saldo_ant",
            "importe_pagado",
            "imp_saldo_insoluto"
        ]
        column_names_str_complemento = ", ".join(column_names_complemento)

        column_names_iva_cobrado_table = [
            "doc_number",
            "account",
            "doc_type",
            "clrng_date",
            "text_iva_cobrado",
            "reference",
            "billing_doc",
            "trading_partner",
            "doc_number_2",
            "clearing_doc",
            "tax_code",
            "doc_date",
            "posting_date",
            "amount_in_local_curr",
            "amount_in_doc_curr",
            "doc_curr",
            "eff_exchange_rate",
            "ano",
            "llave",
            "base_16",
            "cero_nacional",
            "cero_extranjero",
            "iva",
            "ieps",
            "iva_retenido",
            "total_factura",
            "diferencia",
            "ieps_base_iva",
            "nombre",
            "pais",
            "ieps_tasa_6",
            "ieps_tasa_7",
            "total_ieps",
            "diferencia_2"
        ]
        column_names_str_iva_cobrado = ", ".join(column_names_iva_cobrado_table)

        column_names_analisis_iva_cobrado = [
            "doc_number",
            "account",
            "reference",
            "doc_type",
            "doc_number_2",
            "clrng_date",
            "assignment_text",
            "text_analisis_iva_cobrado",
            "clearing_doc",
            "billing_doc",
            "tx",
            "amount_in_local_curr",
            "local_curr",
            "doc_date",
            "posting_date",
            "amount_in_dc",
            "diff",
            "ano",
            "llave",
            "base_16",
            "cero_nacional",
            "cero_extranjero",
            "iva",
            "ieps",
            "iva_retenido",
            "total_factura",
            "diferencia",
            "empty_col_1",
            "ieps_base_iva",
            "nombre",
            "pais",
            "tasa_9",
            "tasa_7",
            "tasa_6",
            "total_ieps",
            "diferencia_2"
        ]
        column_names_str_analisis_iva_cobrado = ", ".join(column_names_analisis_iva_cobrado)

    except Exception as e:
        print(f"Error retrieving column names or constructing query: {e}")
        exit()

    # Prepare data for bulk insertion CFDI Ingreso sheet
    data_collection_ingreso = custom_values_for_insert(data_sheet=cfdi_ingresos_sheet, max_col=21)
    final_result_ingreso_data = ",".join(data_collection_ingreso)
    # Prepare data for bulk insertion CFDI Complemento sheet
    data_collection_complemento = custom_values_for_insert(data_sheet=cfdi_complemento_sheet, max_col=33)
    final_result_complemento_data = ",".join(data_collection_complemento)
    # Prepare data for bulk insertion IVA Cobrado BCS sheet
    data_collection_iva_cobrado_table = custom_values_for_insert(data_sheet=iva_cobrado_sheet, max_col=34)
    final_result_iva_cobrado_data = ",".join(data_collection_iva_cobrado_table)
    # Prepare data for bulk insertion Analisis IVA Cobrado BHC sheet
    data_collection_analisis_iva_cobrado_table = custom_values_for_insert(data_sheet=analisis_iva_cobrado_sheet, max_col=36)
    final_result_analisis_iva_cobrado_data = ",".join(data_collection_analisis_iva_cobrado_table)

    insert_query = f"""SET datestyle = dmy; 
        INSERT INTO cfdi_ingreso ({column_names_str_ingresos}) VALUES {final_result_ingreso_data};
        INSERT INTO complemento ({column_names_str_complemento}) VALUES {final_result_complemento_data};
        INSERT INTO iva_cobrado_bcs ({column_names_str_iva_cobrado}) VALUES {final_result_iva_cobrado_data};
        INSERT INTO analisis_iva_cobrado_bhc ({column_names_str_analisis_iva_cobrado}) VALUES {final_result_analisis_iva_cobrado_data};
    """
    conn, cur = connection_db()
    try:
        # Execute bulk insert using executemany
        execute_query(cur, conn, insert_query)
    except Exception as err:
        print(err)

# def export_data_conciliaciones():
#     start_date = request.args.get("start_date")
#     end_date = request.args.get("end_date")
#     rfc = request.args.get("rfc")
#     # Create a new Excel workbook
#     wb = Workbook()
#     # Add a worksheet
#     ws = wb.active
#     ws.title = f"CFO Data {rfc} from {start_date} to {end_date}"  # Set worksheet title

